# -*- coding: utf-8 -*-

"""解析xlsx数据模块."""

import os
import pickle

from openpyxl import load_workbook

from logger import Logger
from data_structure.data import Data
from data_structure.db import DB


class ParseXlsx:
    """解析xlsx数据类."""

    datapath = f"{os.path.realpath(os.path.dirname(__file__))}/SourceData"

    def __init__(self, filename):
        """解析xlsx数据初始化函数.

        Args:
            filename: 文件路径
              如果filename是绝对路径,则以filename为准
              如果filename是相对路径,则文件需要放在项目根目录下的data目录下
        """
        if not os.path.isabs(filename):
            self._filepath = os.path.join(self.datapath, filename)
        else:
            self._filepath = filename
        if not os.path.exists(self._filepath):
            raise Exception(f"{self._filepath}不存在")
        if not self._filepath.endswith("xlsx"):
            raise Exception("文件需要以xlsx结尾")
        wb = load_workbook(self._filepath)
        self.ws = wb.active
        self._cur_category = None
        self._dbs = dict()
        self.__read_data()

    def __read_data(self):
        for row_idx, row in enumerate(self.ws.iter_rows()):
            self.__parse_row(row_idx, row)

    def __parse_row(self, row_idx, row):
        category_name = row[0].value
        self.__create_data(
            row_idx, row[1].value, row[2].value,
            row[3].value, category_name, row[4].value, row[5].value
        )
        self.__create_data(
            row_idx, row[6].value, row[7].value,
            row[8].value, category_name, row[9].value, row[10].value
        )

    def __create_data(
            self, row_idx, name, datatype, plc_addr, category_name, comment, is_display):
        if None in (name, datatype, plc_addr):
            return
        data = Data(
            name,
            datatype,
            plc_addr,
            category_name,
            comment,
            is_display == 1
        )
        data.row_idx = row_idx
        db = self._dbs.get(data.db_idx)
        if db is None:
            db = DB(data.db_idx)
            self._dbs.update({data.db_idx: db})
        db.add_data(data)

    def iter(self) -> Data:
        for db_idx, db in self._dbs.items():
            for data in db.datas:
                yield db, data

    def print_datas(self):
        for db_idx, db in self._dbs.items():
            for data in db.datas:
                print("\t", data)

    def dump(self, fd):
        for db_idx, db in self._dbs.items():
            pickle.dump(db, fd)


if __name__ == '__main__':
    obj = ParseXlsx("auto-fill-ammo.xlsx")
    obj.print_datas()
