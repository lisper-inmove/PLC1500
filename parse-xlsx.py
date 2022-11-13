# -*- coding: utf-8 -*-

"""解析xlsx数据模块."""

import os

from openpyxl import load_workbook

from logger import Logger
from datatype import Category
from datatype import Data


class ParseXlsx:
    """解析xlsx数据类."""

    datapath = f"{os.path.realpath(os.path.dirname(__file__))}/data"

    def __init__(self, filename):
        """解析xlsx数据初始化函数.

        Args:
            filename: 文件路径
              如果filename是绝对路径,则以filename为准
              如果filename是相对路径,则文件需要放在项目根目录下的data目录下
        """
        if not os.path.isabs(filename):
            self._filepath = os.path.join(self.datapath, filename)
        else:
            self._filepath = filename
        if not os.path.exists(self._filepath):
            raise Exception(f"{self._filepath}不存在")
        if not self._filepath.endswith("xlsx"):
            raise Exception("文件需要以xlsx结尾")
        wb = load_workbook(self._filepath)
        self.ws = wb.active
        self._cur_category = None
        self._categories = dict()
        self.__read_data()

    def __read_data(self):
        for row_idx, row in enumerate(self.ws.iter_rows()):
            if row_idx in [0, 1, 2, 3]: # 忽略头部几行
                continue
            self.__parse_row(row)

    def __parse_row(self, row):
        category_name = row[0].value
        if category_name is not None:
            self._cur_category = Category(category_name)
            self._categories.update({category_name: self._cur_category})
        if None not in [row[1].value, row[2].value, row[3].value]:
            self._cur_category.add_data(
                Data(row[1].value, row[2].value, row[3].value, row[4].value)
            )
        if None not in [row[5].value, row[6].value, row[7].value]:
            self._cur_category.add_data(
                Data(row[5].value, row[6].value, row[7].value, row[8].value)
            )

    def print_datas(self):
        for category_name, category in self._categories.items():
            print(category_name)
            for data in category.datas:
                print("\t", data)


if __name__ == '__main__':
    obj = ParseXlsx("S-1500-data.xlsx")
    obj.print_datas()
